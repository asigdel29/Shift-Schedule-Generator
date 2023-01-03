import openpyxl

wb = openpyxl.load_workbook('Spring2023.xlsx')
availability_sheet = wb['Availability']
schedule_sheet = wb['Schedule']

employee_availability = {}

for row in availability_sheet.iter_rows(min_row=2, max_row=15, min_col=2, max_col=13):
    # Get the time of the shift
    time = row[0].value

    for cell in row[1:]:
        name = cell.value

        if name not in employee_availability:
            employee_availability[name] = {}

        employee_availability[name][time] = True

employees = list(employee_availability.keys())
employee_work_hours = {employee: 0 for employee in employees}
unfilled_shifts = []

for row in schedule_sheet.iter_rows(min_row=2, max_row=15, min_col=2, max_col=13):
    time = row[0].value

    for cell in row[1:]:
        if cell.value is None:
            unfilled_shifts.append((time, cell))

unfilled_shifts.sort(key=lambda shift: employee_work_hours[shift[1].value])

for shift in unfilled_shifts:
    time, cell = shift

    for employee in employees:
        # If the employee is available for the shift and has fewer work hours than the other employees, assign them to the shift
        if employee_availability[employee].get(time, False) and employee_work_hours[employee] <= min(employee_work_hours.values()):
            cell.value = employee
            employee_work_hours[employee] += 1
            break

wb.save('Spring2023.xlsx')
